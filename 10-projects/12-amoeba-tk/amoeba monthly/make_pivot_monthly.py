import sys
sys.stdout.reconfigure(encoding='utf-8')
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 파일 경로 ────────────────────────────────────────────
FILE_IN  = 'C:/iloom-workspace/10-projects/12-amoeba-tk/amoeba monthly/02. 분개장정보상세(3월가)_처리완료.xlsx'
FILE_OUT = FILE_IN

df = pd.read_excel(FILE_IN, sheet_name='RAW')
print(f"RAW 읽기 완료: {len(df)}행  컬럼: {list(df.columns)}")

팀순서       = ['리테일', '직영', '온라인']
직접간접순서 = ['간접', '직접']

# ── 피봇 계산 ────────────────────────────────────────────
pivot = df.groupby(
    ['채산계정1','채산계정2','채산계정3','세부','직접간접']
)['금액(안분후)'].sum().reset_index()

def get_val(c1, c2, c3, 팀, di):
    rows = pivot[
        (pivot['채산계정1']==c1) &
        (pivot['채산계정2']==c2) &
        (pivot['채산계정3']==c3) &
        (pivot['세부']==팀) &
        (pivot['직접간접']==di)
    ]
    return int(rows['금액(안분후)'].sum()) if len(rows) > 0 else 0

def get_c1total(c1, 팀, di):
    rows = pivot[
        (pivot['채산계정1']==c1) &
        (pivot['세부']==팀) &
        (pivot['직접간접']==di)
    ]
    return int(rows['금액(안분후)'].sum()) if len(rows) > 0 else 0

# ── 계층 구조 ────────────────────────────────────────────
계층 = {
    '고정비': {
        '일반관리비': sorted(['(판)견본비','(판)경상연구개발비','(판)교육훈련비','(판)도서인쇄비',
                       '(판)무형자산감가상각비','(판)물류비(AS)','(판)물류비(재고보관)',
                       '(판)보험료','(판)복리후생비','(판)세금과공과','(판)소모품비',
                       '(판)수도광열비','(판)수선비','(판)여비교통비','(판)유형자산감가상각비',
                       '(판)임차료','(판)접대비','(판)지급수수료','(판)차량유지비',
                       '(판)통신비','(판)하자보수비','(판)비품비']),
        '마케팅비': ['(판)광고선전비','(판)판촉비'],
    },
    '변동비': {
        '물류비': ['(판)물류비(물류)_바로스','(판)물류비(물류)_시디즈','(판)물류비(시공)',
                  '(판)물류비(운송)_바로스','(판)물류비(운송)_시디즈'],
        '(판)판매수수료': ['(판)판매수수료_B/S','(판)판매수수료_PG사','(판)판매수수료_PSA',
                         '(판)판매수수료_기타','(판)판매수수료_외부몰','(판)판매수수료_입점몰',
                         '(판)판매수수료_투자B'],
    },
    '영업외손익': {
        '수입수수료':          ['수입수수료'],
        '잡이익':              ['잡이익'],
        '잡손실':              ['잡손실'],
        '임대료':              ['임대료'],
        '임차료(영업외)':      ['임차료(영업외)'],
        '유형자산처분이익':    ['유형자산처분이익'],
        '지급수수료(영업외)':  ['지급수수료(영업외)'],
    },
}

# ── 스타일 ───────────────────────────────────────────────
COLOR = {
    '고정비':     'FF92D050',
    '변동비':     'FF00B0F0',
    '영업외손익': 'FFFF0000',
    'header':     'FF404040',
    'header_sub': 'FF595959',
    'total':      'FFFFF2CC',
}

def make_fill(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color)

white_font = Font(name='맑은 고딕', bold=True, color='FFFFFFFF', size=9)
black_font = Font(name='맑은 고딕', size=9)
bold_font  = Font(name='맑은 고딕', bold=True, size=9)
center_al  = Alignment(horizontal='center', vertical='center', wrap_text=True)
right_al   = Alignment(horizontal='right',  vertical='center')

# ── 시트 생성 ────────────────────────────────────────────
wb = openpyxl.load_workbook(FILE_IN)

if '채산피봇' in wb.sheetnames:
    del wb['채산피봇']

raw_idx = wb.sheetnames.index('RAW')
ws = wb.create_sheet('채산피봇', raw_idx + 1)

COL_LABEL  = 1
DATA_START = 2

def team_col_start(team_idx):
    return DATA_START + team_idx * 3

TOTAL_COL = DATA_START + len(팀순서) * 3

# ── 행1: 팀 헤더 ─────────────────────────────────────────
ws.cell(1, COL_LABEL).value = '행 레이블'
ws.cell(1, COL_LABEL).font  = white_font
ws.cell(1, COL_LABEL).fill  = make_fill(COLOR['header'])
ws.cell(1, COL_LABEL).alignment = center_al

for i, 팀 in enumerate(팀순서):
    cs = team_col_start(i)
    ws.cell(1, cs).value = f'={팀}'
    ws.cell(1, cs).font  = white_font
    ws.cell(1, cs).fill  = make_fill(COLOR['header'])
    ws.cell(1, cs).alignment = center_al
    ws.merge_cells(start_row=1, start_column=cs, end_row=1, end_column=cs+1)
    ws.cell(1, cs+2).value = f'{팀} 요약'
    ws.cell(1, cs+2).font  = white_font
    ws.cell(1, cs+2).fill  = make_fill(COLOR['header'])
    ws.cell(1, cs+2).alignment = center_al

ws.cell(1, TOTAL_COL).value = '총합계'
ws.cell(1, TOTAL_COL).font  = white_font
ws.cell(1, TOTAL_COL).fill  = make_fill(COLOR['header'])
ws.cell(1, TOTAL_COL).alignment = center_al

# ── 행2: 간접/직접 헤더 ─────────────────────────────────
ws.cell(2, COL_LABEL).fill = make_fill(COLOR['header'])
for i, 팀 in enumerate(팀순서):
    cs = team_col_start(i)
    for j, di in enumerate(직접간접순서):
        ws.cell(2, cs+j).value = di
        ws.cell(2, cs+j).font  = white_font
        ws.cell(2, cs+j).fill  = make_fill(COLOR['header_sub'])
        ws.cell(2, cs+j).alignment = center_al
    ws.cell(2, cs+2).fill = make_fill(COLOR['header'])

ws.cell(2, TOTAL_COL).fill = make_fill(COLOR['header'])
ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 18

current_row = 3

def write_row(row, label, vals, fill_color=None, font=None):
    ws.cell(row, COL_LABEL).value = label
    if fill_color:
        ws.cell(row, COL_LABEL).fill = make_fill(fill_color)
    ws.cell(row, COL_LABEL).font = font or black_font
    ws.cell(row, COL_LABEL).alignment = Alignment(vertical='center', indent=1)

    for i, 팀 in enumerate(팀순서):
        cs = team_col_start(i)
        for j, di in enumerate(직접간접순서):
            v = vals.get((팀, di), 0)
            c = ws.cell(row, cs+j)
            c.value = v if v != 0 else None
            c.number_format = '#,##0'
            c.alignment = right_al
            c.font = font or black_font
            if fill_color:
                c.fill = make_fill(fill_color)
        subtotal = sum(vals.get((팀, di), 0) for di in 직접간접순서)
        sc = ws.cell(row, cs+2)
        sc.value = subtotal if subtotal != 0 else None
        sc.number_format = '#,##0'
        sc.alignment = right_al
        sc.font = font or black_font
        if fill_color:
            sc.fill = make_fill(fill_color)

    total = sum(vals.get((팀, di), 0) for 팀 in 팀순서 for di in 직접간접순서)
    tc = ws.cell(row, TOTAL_COL)
    tc.value = total if total != 0 else None
    tc.number_format = '#,##0'
    tc.alignment = right_al
    tc.font = font or black_font
    if fill_color:
        tc.fill = make_fill(fill_color)

# ── 데이터 행 출력 ────────────────────────────────────────
for c1 in ['고정비', '변동비', '영업외손익']:
    color = COLOR[c1]
    c1_vals = {(팀, di): get_c1total(c1, 팀, di) for 팀 in 팀순서 for di in 직접간접순서}
    write_row(current_row, f'={c1}', c1_vals, fill_color=color, font=white_font)
    ws.row_dimensions[current_row].height = 16
    current_row += 1

    for c2, c3_list in 계층[c1].items():
        for c3 in c3_list:
            vals = {(팀, di): get_val(c1, c2, c3, 팀, di) for 팀 in 팀순서 for di in 직접간접순서}
            write_row(current_row, c3, vals)
            ws.row_dimensions[current_row].height = 14
            current_row += 1

# ── 총합계 행 ────────────────────────────────────────────
total_vals = {
    (팀, di): int(pivot[(pivot['세부']==팀) & (pivot['직접간접']==di)]['금액(안분후)'].sum())
    for 팀 in 팀순서 for di in 직접간접순서
}
write_row(current_row, '총합계', total_vals, fill_color=COLOR['total'], font=bold_font)
ws.row_dimensions[current_row].height = 18
current_row += 1

# ── 열 너비 & 고정 ───────────────────────────────────────
ws.column_dimensions[get_column_letter(COL_LABEL)].width = 28
for i in range(len(팀순서)):
    cs = team_col_start(i)
    for j in range(3):
        ws.column_dimensions[get_column_letter(cs+j)].width = 16
ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 16

ws.freeze_panes = 'B3'

wb.save(FILE_OUT)
print(f'완료! 채산피봇 시트 생성됨. 총 {current_row-3}행')
