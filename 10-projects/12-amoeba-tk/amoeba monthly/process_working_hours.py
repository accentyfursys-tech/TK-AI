"""
근무시간상세 파일 처리 스크립트
1. 공통직접구분 열 옆에 '직접간접' 열 추가 (빈칸)
2. 안분된채산조직 괄호 값을 별도 행으로 분리 (단, '일룸' 제외)
"""

import re
import openpyxl
from copy import copy

INPUT_FILE = "03. 근무시간상세(2월찐).xlsx"
OUTPUT_FILE = "03. 근무시간상세(2월찐)_처리완료.xlsx"

# 공통직접구분: 5번째 컬럼(col index 5, 1-based)
# 안분된채산조직: 7번째 컬럼(col index 7, 1-based)
COL_공통직접 = 5   # E열


def parse_org(value):
    """
    조직명에서 기본명과 괄호값 분리, '일룸' 제외
    예: '온라인CX개선팀(공식몰)(일룸)' → base='온라인CX개선팀', 팀세부='공식몰'
    예: '일룸직영(강동아이파크)'        → base='일룸직영',      팀세부='강동아이파크'
    """
    if not value:
        return value, None
    base = re.sub(r'\([^)]*\)', '', str(value)).strip()
    brackets = [b for b in re.findall(r'\(([^)]+)\)', str(value)) if b != '일룸']
    detail = brackets[0] if brackets else None
    return base, detail


def main():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active

    # ── Step 1: E열(공통직접구분) 오른쪽에 '직접간접' 열 삽입 ──
    col_직접간접 = COL_공통직접 + 1  # F열
    ws.insert_cols(col_직접간접)
    ws.cell(1, col_직접간접).value = "직접간접"
    # 삽입 후 안분된채산조직: 원래 7번 → 8번
    col_안분 = 8

    print("'직접간접' 열 삽입 완료 (F열)")

    # ── Step 2: 안분된채산조직 옆에 '팀세부' 열 삽입 ──
    col_팀세부 = col_안분 + 1  # 9번
    ws.insert_cols(col_팀세부)
    ws.cell(1, col_팀세부).value = "팀세부"

    print("'팀세부' 열 삽입 완료 (I열)")

    # ── Step 3: 각 행 파싱 → 안분된채산조직=기본명, 팀세부=괄호값 ──
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, col_안분).value
        base, detail = parse_org(val)
        ws.cell(row_idx, col_안분).value = base
        ws.cell(row_idx, col_팀세부).value = detail

    print(f"안분된채산조직 파싱 완료 ({ws.max_row - 1}행)")

    # ── Step 4: '부서' 열 삽입 (안분된채산조직 뒤, 팀세부 뒤 각각) ──
    # 안분된채산조직(8) 뒤에 부서 삽입 → 9번, 팀세부는 10번으로 밀림
    col_부서1 = col_안분 + 1   # 9번 (안분된채산조직 옆)
    ws.insert_cols(col_부서1)
    ws.cell(1, col_부서1).value = "부서"
    # 팀세부는 10번으로 밀림
    col_팀세부 = 10
    # 팀세부 뒤에 부서 삽입 → 11번
    col_부서2 = col_팀세부 + 1  # 11번
    ws.insert_cols(col_부서2)
    ws.cell(1, col_부서2).value = "부서"

    # 안분된채산조직 → 부서 매핑
    안분_매핑 = {
        '온라인CX개선팀': '온라인',
        '일룸직영': '직영',
        '일룸유통': '리테일',
    }
    # 팀세부 → 부서 매핑
    팀세부_매핑 = {
        '공식몰': '온라인',
        '외부몰': '온라인',
        '강동아이파크': '직영',
        '노원': '직영',
        '논현': '직영',
        '대구': '직영',
        '대전둔산': '직영',
        '마포서대문': '직영',
        '부산센텀': '직영',
        '분당서현': '직영',
        '송파': '직영',
        '수원광교': '직영',
        '용산': '직영',
        '기타': '직영',
        '일반': '리테일',
        '투자': '리테일',
    }

    for row_idx in range(2, ws.max_row + 1):
        v_안분 = ws.cell(row_idx, col_안분).value
        v_팀세부 = ws.cell(row_idx, col_팀세부).value
        ws.cell(row_idx, col_부서1).value = 안분_매핑.get(v_안분, None)
        ws.cell(row_idx, col_부서2).value = 팀세부_매핑.get(v_팀세부, None)

    print("'부서' 열 입력 완료")

    # ── Step 5: '직접간접' 열 데이터 입력 ──
    # 코스트센터 컬럼 위치 확인
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_cc = header.index('코스트센터') + 1
    col_직접간접 = header.index('직접간접') + 1

    직접_키워드 = ['일룸유통', '유통경쟁력강화팀', '온라인CX개선팀', '리테일사업팀']

    for row_idx in range(2, ws.max_row + 1):
        cc = ws.cell(row_idx, col_cc).value or ''
        if any(k in cc for k in 직접_키워드):
            ws.cell(row_idx, col_직접간접).value = '직접'
        else:
            ws.cell(row_idx, col_직접간접).value = '간접'

    print("'직접간접' 열 입력 완료")

    # ── 저장 ──
    wb.save(OUTPUT_FILE)
    print(f"\n저장 완료: {OUTPUT_FILE}")
    print(f"최종 행 수: {ws.max_row - 1}개 (헤더 제외)")


if __name__ == "__main__":
    main()
