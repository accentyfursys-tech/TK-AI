# -*- coding: utf-8 -*-
"""
01. 상제품매출상세 (2월찐).xlsx RAW 시트 변환 스크립트

변환 내용:
1. E열에 "매출구분" 삽입 (채산계정 값 복사)
2. 실적사업소 다음 열에 "채널구분" 삽입
3. 안분된채산조직_계층을 // 기준으로 3열 세분화
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import copy

print("파일 로딩 중...")
df = pd.read_excel('01. 상제품매출상세 (2월찐).xlsx', sheet_name='RAW')
print(f"로딩 완료: {len(df)}행 x {len(df.columns)}열")

# ============================
# 1. E열(인덱스 4) 앞에 "매출구분" 열 삽입
# ============================
# 현재 컬럼 순서: INDEX(), 회사, 사업부, 매출일자, 건명, ...
# E열(5번째)에 매출구분 삽입 → 건명이 F열로 밀림
매출구분 = df['채산계정'].copy()

cols = list(df.columns)
insert_pos_e = 4  # 0-based, 5번째 열 앞
df.insert(insert_pos_e, '매출구분', 매출구분)
print(f"매출구분 열 삽입 완료 (E열)")

# ============================
# 2. 실적사업소 다음에 "채널구분" 삽입
# ============================
def get_채널구분(채산조직):
    if pd.isna(채산조직):
        return ''
    s = str(채산조직)
    if '직영' in s:
        return '직영'
    elif '온라인CX개선팀' in s and '공식몰' in s:
        return '온라인(공식몰)'
    elif '온라인CX개선팀' in s and '외부몰' in s:
        return '온라인(외부몰)'
    elif '유통' in s:
        return '유통'
    else:
        return s

채널구분_values = df['안분된채산조직'].apply(get_채널구분)

실적사업소_pos = df.columns.get_loc('실적사업소')
df.insert(실적사업소_pos + 1, '채널구분', 채널구분_values)
print(f"채널구분 열 삽입 완료 (실적사업소 다음)")

# ============================
# 3. 안분된채산조직_계층 세분화 (// 기준 3열 분리)
# ============================
def split_hierarchy(val):
    if pd.isna(val):
        return '', '', ''
    parts = str(val).split('//')
    l1 = parts[0].strip() if len(parts) > 0 else ''
    l2 = parts[1].strip() if len(parts) > 1 else ''
    l3 = parts[2].strip() if len(parts) > 2 else ''
    return l1, l2, l3

hierarchy_split = df['안분된채산조직_계층'].apply(split_hierarchy)
df['채산조직_사업부'] = [x[0] for x in hierarchy_split]
df['채산조직_팀'] = [x[1] for x in hierarchy_split]
df['채산조직_채널'] = [x[2] for x in hierarchy_split]

# 안분된채산조직_계층 다음에 3열 삽입 (현재 위치 유지하며 뒤에 추가)
계층_pos = df.columns.get_loc('안분된채산조직_계층')
# 현재 df에서 이미 뒤에 추가됨 → 계층 바로 다음으로 재배치
cols_now = list(df.columns)
new_cols = []
for c in cols_now:
    new_cols.append(c)
    if c == '안분된채산조직_계층':
        new_cols += ['채산조직_사업부', '채산조직_팀', '채산조직_채널']

# 중복 제거 (이미 추가된 것들 제외)
seen = set()
final_cols = []
for c in new_cols:
    if c not in seen:
        final_cols.append(c)
        seen.add(c)

df = df[final_cols]
print(f"안분된채산조직_계층 세분화 완료")

# ============================
# 결과 저장
# ============================
print(f"\n최종 컬럼 수: {len(df.columns)}")
for i, c in enumerate(df.columns):
    print(f"  {get_column_letter(i+1)}열: {c}")

output_file = '01. 상제품매출상세 (2월찐).xlsx'
print(f"\n저장 중: {output_file}")

# 기존 파일에 RAW 시트 덮어쓰기
with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    df.to_excel(writer, sheet_name='RAW', index=False)

print("저장 완료!")
print(f"최종: {len(df)}행 x {len(df.columns)}열")
